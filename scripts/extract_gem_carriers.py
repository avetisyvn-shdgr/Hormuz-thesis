"""Extract the frozen GEM LNG carrier workbook deterministically."""
from __future__ import annotations

import argparse
from datetime import date, datetime, time
import hashlib
import json
import math
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "data/raw/gem/LNG-Carrier-Tracker-December-2025-release.xlsx"
DEFAULT_OUTPUT = ROOT / "data/interim/gem_lng_carriers.json"
DEFAULT_METADATA_OUTPUT = ROOT / "data/interim/gem_lng_carriers_metadata.json"
DEFAULT_SHEET = "data"
REQUIRED_COLUMNS = {
    "IMO number",
    "IMO number [ref]",
    "Name",
    "Status",
    "Capacity",
    "Capacity [ref]",
    "Vessel type",
    "Delivery year",
    "Shipowner",
    "Propulsion type",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _relative_or_absolute(path: Path) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(ROOT).as_posix()
    except ValueError:
        return resolved.as_posix()


def _normalize_cell(value: object, *, epoch: datetime) -> object:
    """Match the numeric representation used by the original workbook export."""
    if isinstance(value, (datetime, date, time)):
        value = to_excel(value, epoch=epoch)
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return int(value)
    return value


def _valid_imo(value: object) -> bool:
    try:
        text = str(int(value))
    except (TypeError, ValueError, OverflowError):
        return False
    if len(text) != 7 or not text.isdigit():
        return False
    checksum = sum(
        int(digit) * weight
        for digit, weight in zip(text[:6], range(7, 1, -1))
    ) % 10
    return checksum == int(text[-1])


def extract(
    input_path: Path,
    output_path: Path,
    metadata_output_path: Path,
    *,
    sheet_name: str = DEFAULT_SHEET,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"worksheet not found: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError(f"worksheet is empty: {sheet_name}") from exc

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    if any(not header for header in headers):
        raise ValueError("worksheet contains an empty column header")
    if len(headers) != len(set(headers)):
        raise ValueError("worksheet contains duplicate column headers")
    missing = sorted(REQUIRED_COLUMNS.difference(headers))
    if missing:
        raise ValueError(f"worksheet is missing required columns: {missing}")

    records: list[dict[str, object]] = []
    for row in rows:
        if not row or row[0] in (None, ""):
            continue
        padded = tuple(row) + (None,) * (len(headers) - len(row))
        records.append({
            header: _normalize_cell(padded[index], epoch=workbook.epoch)
            for index, header in enumerate(headers)
        })

    imos = [record["IMO number"] for record in records]
    normalized_imos = [str(int(value)) for value in imos if _valid_imo(value)]
    if len(normalized_imos) != len(records):
        raise ValueError("worksheet contains an invalid IMO number")
    if len(normalized_imos) != len(set(normalized_imos)):
        raise ValueError("worksheet contains duplicate IMO numbers")

    metadata: dict[str, object] = {
        "schema_version": 1,
        "source_file": _relative_or_absolute(input_path),
        "source_sha256": _sha256(input_path),
        "source_sheet": sheet_name,
        "extracted_rows": len(records),
        "extracted_columns": headers,
        "primary_key": "IMO number",
        "primary_key_validation": "IMO check digit and uniqueness",
        "date_encoding": "Excel serial numbers preserved from the source workbook",
        "extraction_tool": "openpyxl",
        "extraction_tool_version": openpyxl.__version__,
        "producer": "scripts/extract_gem_carriers.py",
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    metadata_output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(records, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    metadata_output_path.write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return records, metadata


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--metadata-output",
        type=Path,
        default=DEFAULT_METADATA_OUTPUT,
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    args = parser.parse_args()

    records, metadata = extract(
        args.input,
        args.output,
        args.metadata_output,
        sheet_name=args.sheet,
    )
    print(f"wrote {args.output}")
    print(f"wrote {args.metadata_output}")
    print(f"rows={len(records)} columns={len(metadata['extracted_columns'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
